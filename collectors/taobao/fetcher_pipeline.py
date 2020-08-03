# -*- coding: utf-8 -*-
import json
import random
import time

import openpyxl
import requests

"""
-------------------------------------------------
   File Name：     fetcher_pipeline
   Description :
   Author :        patrick
   date：          2019/11/20
-------------------------------------------------
   Change Activity:
                   2019/11/20:
-------------------------------------------------
"""
__author__ = 'patrick'

session = requests.session()
headers = {
    'accept': '*/*',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'zh-CN,zh;q=0.9',
    'referer': 'https://uland.taobao.com/semm/tbsearch?refpid=mm_26632258_3504122_32554087&keyword=%E5%A5%B3%E8'
               '%A3%85 '
               '&rewriteQuery=1&a=mi={imei}&sms=baidu&idfa={'
               'idfa}&clk1=abab6283306413775910d4b0b37ca047&upsid=abab6283306413775910d4b0b37ca047',
    'user-agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/%s Mobile Safari/537.36',
}


def fetch_data(page, key_word):
    url = 'https://odin.re.taobao.com/m/Nwalltbuad?sbid=sem2_kgb_activity&ignore=CATID%2CRANKINFO%2CMATCHTYPE&pvid=_TL' \
          '-41832&refpid=mm_26632258_3504122_32554087&clk1=abab6283306413775910d4b0b37ca047&idfa=%7Bidfa%7D&pid' \
          '=430680_1006&keyword=' + key_word + '&count=60&offset=' + str(60 * page) + '&relacount=8&t=1535075213992' \
                                                                                     '&callback' \
                                                                                     '=mn17jsonp1535075213992'
    r = session.get(url=url, headers=headers)
    html = r.text
    start = html.find('(')
    return (json.loads(html[start + 1:-1]))['result']['item']


def save_to_excel(feed_data, file_name="item.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row = 1
    for p in feed_data:
        # Put the key in the first column for each key in the dictionary
        column = 1
        for key, value in p.items():
            sheet.cell(row=row, column=column, value=key)
            sheet.cell(row=row, column=column + 1, value=value)
            column += 2
        row += 1

    workbook.save(filename=file_name + ".xlsx")


if __name__ == '__main__':
    result = []
    keyword = "pvdc保鲜膜".replace(" ", "+")
    for i in range(10):
        data = fetch_data(i+1, keyword)
        result.extend(data)
        time.sleep(random.choice([5, 6, 7, 8, 9, 10]))
    save_to_excel(result, keyword)
